import os
import glob
import pandas as pd
from copy import copy
from openpyxl import load_workbook, Workbook
from config import *
from logger import get_logger

logger = get_logger(__name__)

def copy_sheet(src_ws, tgt_wb, title):
    """
    src_ws 워크시트의 내용(값, 수식, 스타일, 병합, 크기 등)을
    tgt_wb 에 'title' 이름의 새 시트로 복사합니다.
    """
    # 중복 방지: 이미 같은 이름이 있으면 _1, _2 붙이기
    final_title = title
    if final_title in tgt_wb.sheetnames:
        base = final_title
        i = 1
        while f"{base}_{i}" in tgt_wb.sheetnames:
            i += 1
        final_title = f"{base}_{i}"

    tgt_ws = tgt_wb.create_sheet(title=final_title)

    # 셀 복사
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = tgt_ws.cell(
                row=cell.row,
                column=cell.col_idx,
                value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            if cell.hyperlink:
                new_cell._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                new_cell.comment = copy(cell.comment)

    # merged cells 복사
    for merged_cell_range in src_ws.merged_cells.ranges:
        tgt_ws.merge_cells(str(merged_cell_range))

    # 열 너비, 행 높이 복사
    for col, dim in src_ws.column_dimensions.items():
        tgt_ws.column_dimensions[col].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        tgt_ws.row_dimensions[row_idx].height = dim.height


#20250724 기존 코드 개선
def merge_with_pandas(input_dir,output_file):
    files=glob.glob(os.path.join(input_dir,"*.xlsx"))

    if not files:
        print("No Excel files found in", input_dir)
        return
    
    with pd.ExcelWriter(output_file,engine='openpyxl') as writer:
        for file in files:
            print(f"file {file} start")
            sheetname= os.path.splitext(os.path.basename(file))[0][:31]
            df = pd.read_excel(file, engine='openpyxl') # 첫 시트만 로딩
            df.to_excel(writer, sheet_name=sheetname, index=False)
            print(f"file {file} completed!")
        
        print(f"Merged {len(files)} files into {output_file}")


def merge_csv_files(input_dir,output_file):
    files = glob.glob(os.path.join(input_dir, "*.csv"))

    if not files:
        print("No csv files found in", input_dir)
        return

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for file in files:
            print(f"file {file} start")
            sheetname= os.path.splitext(os.path.basename(file))[0][:31]
            df = pd.read_csv(file, encoding='utf-8-sig', low_memory=False)
            df.to_excel(writer,sheet_name=sheetname, index=False)
            print(f"file {file} completed!")

        print(f"Merged {len(files)} files into {output_file}")
        
    delete_all_file(input_dir)

def merge_excels_preserve(input_dir, output_dir):
    pattern = os.path.join(input_dir, "*.xlsx")
    output_file = os.path.join(output_dir,'merge.xlsx')
    excel_files = glob.glob(pattern)

    if not excel_files:
        logger.info(f"No Excel files found in : {input_dir}")
        return

    out_wb = Workbook()
    out_wb.remove(out_wb.active)  # 기본 시트 제거

    logger.info(f"=== Start to merge files===")
    for file in excel_files:
        # 파일명(확장자 제외)을 시트명으로 사용 (최대 31자)
        base_name = os.path.splitext(os.path.basename(file))[0][:31]

        # 워크북 로드: data_only=False → 수식(formula) 그대로 읽기
        src_wb = load_workbook(file, data_only=False)
        # 첫 번째 시트만 복사
        src_ws = src_wb.worksheets[0]
        copy_sheet(src_ws, out_wb, base_name)
        logger.info(f"Copy finished : {base_name} ")

    out_wb.save(output_file)
    logger.info(f"Merged {len(excel_files)} files into {output_file}")
    
def delete_all_file(dir : str):
    for filename in os.listdir(dir):
        file_path = os.path.join(dir, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)